# soc_importer.py
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (TimeoutException, NoSuchWindowException)
import openpyxl as xl
import configparser
import time
import logging

from base_web_bot import BaseWebBot
from soc_base_mixin import SOC_BaseMixin


class SOC_Importer(BaseWebBot, SOC_BaseMixin):
    """Specialized bot for importing SOC overrides from Excel files."""

    def __init__(self, soc_id=None):
        BaseWebBot.__init__(self)
        SOC_BaseMixin.__init__(self)
        self.load_configuration()
        self.SOC_base_link = self.base_link + r"Soc/EditOverrides/"
        self.SOC_id = soc_id

    @property
    def base_link(self) -> str:
        return self._base_link        

    def load_configuration(self):
        config = configparser.ConfigParser(interpolation=None)
        config.read(self.config_file, encoding="utf8")

        self.user_name = config.get('Settings', 'user_name', fallback='xxxxxx')
        raw_password = config.get('Settings', 'password', fallback='******')
        self.password = self.process_password(raw_password)

        if '\n' in self.password:
            self.password = 'INCORRECT PASSWORD'

        self._base_link = config.get('Settings', 'base_link', fallback='http://eptw.sakhalinenergy.ru/')
        self.MAX_WAIT_PAGE_LOAD_DELAY_SECONDS = config.getint('Settings', 'MAX_WAIT_PAGE_LOAD_DELAY_SECONDS', fallback=20)

    def load_overrides_from_export(self):
        """Load overrides from Excel file."""
        try:
            wb = xl.load_workbook('soc_import_export/overrides.xlsx')
            sheet = wb.active

            columns = ['TagNumber', 'Description', 'OverrideType', 'OverrideMethod', 
                      'Comment', 'AppliedState', 'AdditionalValueAppliedState', 
                      'RemovedState', 'AdditionalValueRemovedState']

            self.list_of_overrides = []
            for row in range(7, sheet.max_row + 1):
                tag_number = sheet.cell(row, 1).value
                if not tag_number:
                    continue

                override = {col: sheet.cell(row, idx + 1).value for idx, col in enumerate(columns)}
                
                for field in ['Comment', 'AdditionalValueAppliedState', 'AdditionalValueRemovedState', 'RemovedState']:
                    if override[field] == "":
                        override[field] = None

                self.list_of_overrides.append(override)

            logging.info(f"✅ Loaded {len(self.list_of_overrides)} overrides")

        except Exception as e:
            logging.error(f"❌ Failed to load overrides: {e}")
            raise

    def execute_script(self, script: str, *args):
        """Execute JavaScript."""
        try:
            return self.driver.execute_script(script, *args)
        except Exception as e:
            logging.warning(f"JavaScript execution failed: {e}")
            return None

    def wait_for_kendo_dropdown(self, element_id: str, timeout: int = 10) -> bool:
        """Wait for Kendo dropdown to be initialized."""
        try:
            WebDriverWait(self.driver, timeout).until(
                lambda driver: self.execute_script(f"return $('#{element_id}').data('kendoDropDownList') !== undefined;")
            )
            return True
        except TimeoutException:
            logging.error(f"❌ Kendo dropdown {element_id} not initialized within {timeout}s")
            return False

    def inject_cascade_script(self):
        """Inject the cascade script."""
        cascade_script = """
window.fixedCascades = {
    typeToMethod: function(typeValue) {
        const ddType = $('#OverrideTypeId').data('kendoDropDownList');
        const ddMethod = $('#OverrideMethodId').data('kendoDropDownList');
        const typeElement = $('#OverrideTypeId')[0];
        
        return new Promise((resolve) => {
            if (!ddType || !ddMethod) {
                resolve(0);
                return;
            }

            const typeEvents = $._data(typeElement, 'events');
            if (!typeEvents?.change?.[0]) {
                resolve(0);
                return;
            }

            const changeHandler = typeEvents.change[0].handler;
            
            try {
                ddType.value(typeValue);
                
                const event = $.Event('change');
                Object.assign(event, {
                    target: typeElement,
                    currentTarget: typeElement,
                    sender: ddType,
                    value: typeValue
                });
                
                changeHandler.call(typeElement, event);
                
                setTimeout(() => {
                    const itemCount = ddMethod.dataItems().length;
                    resolve(itemCount);
                }, 800);
                
            } catch (error) {
                resolve(0);
            }
        });
    },
    
    methodToStates: function(methodValue) {
        const ddMethod = $('#OverrideMethodId').data('kendoDropDownList');
        const ddApplied = $('#OverrideAppliedStateId').data('kendoDropDownList');
        const ddRemoved = $('#OverrideRemovedStateId').data('kendoDropDownList');
        const methodElement = $('#OverrideMethodId')[0];
        
        return new Promise((resolve) => {
            if (!ddMethod || !ddApplied || !ddRemoved) {
                resolve({ applied: 0, removed: 0 });
                return;
            }

            const methodEvents = $._data(methodElement, 'events');
            if (!methodEvents?.change?.[0]) {
                resolve({ applied: 0, removed: 0 });
                return;
            }

            const changeHandler = methodEvents.change[0].handler;
            
            try {
                const numericValue = parseInt(methodValue);
                ddMethod.value(numericValue);
                
                const event = $.Event('change');
                Object.assign(event, {
                    target: methodElement,
                    currentTarget: methodElement,
                    sender: ddMethod,
                    value: numericValue
                });
                
                changeHandler.call(methodElement, event);
                
                setTimeout(() => {
                    const appliedCount = ddApplied.dataItems().length;
                    const removedCount = ddRemoved.dataItems().length;
                    resolve({ applied: appliedCount, removed: removedCount });
                }, 800);
                
            } catch (error) {
                resolve({ applied: 0, removed: 0 });
            }
        });
    }
};
"""
        try:
            self.execute_script(cascade_script)
            time.sleep(0.5)
            return self.execute_script("return typeof fixedCascades !== 'undefined';")
        except Exception as e:
            logging.error(f"❌ Failed to inject cascade script: {e}")
            return False

    def cascade_type_to_method(self, type_value: int) -> bool:
        """Cascade Type selection to Method dropdown."""
        try:
            if not self.inject_cascade_script():
                return False
            
            if not (self.wait_for_kendo_dropdown("OverrideTypeId") and 
                    self.wait_for_kendo_dropdown("OverrideMethodId")):
                return False

            result = self.execute_script("""
                return fixedCascades.typeToMethod(arguments[0]).then(function(count) {
                    return count > 0;
                });
            """, type_value)
            
            time.sleep(1)
            
            method_count = self.execute_script("""
                var ddMethod = $('#OverrideMethodId').data('kendoDropDownList');
                return ddMethod ? ddMethod.dataItems().length : 0;
            """)
            
            success = method_count and method_count > 0
            if success:
                logging.info(f"✅ Type {type_value} → {method_count} methods")
            else:
                logging.error(f"❌ Type cascade failed")
                
            return success
            
        except Exception as e:
            logging.error(f"❌ Type cascade failed: {e}")
            return False

    def cascade_method_to_states(self, method_value: int) -> bool:
        """Cascade Method selection to State dropdowns."""
        try:
            if not self.inject_cascade_script():
                return False
            
            if not (self.wait_for_kendo_dropdown("OverrideMethodId") and 
                    self.wait_for_kendo_dropdown("OverrideAppliedStateId") and
                    self.wait_for_kendo_dropdown("OverrideRemovedStateId")):
                return False

            result = self.execute_script("""
                return fixedCascades.methodToStates(arguments[0]).then(function(counts) {
                    return counts.applied > 0 && counts.removed > 0;
                });
            """, method_value)
            
            time.sleep(1)
            
            state_counts = self.execute_script("""
                var ddApplied = $('#OverrideAppliedStateId').data('kendoDropDownList');
                var ddRemoved = $('#OverrideRemovedStateId').data('kendoDropDownList');
                return {
                    applied: ddApplied ? ddApplied.dataItems().length : 0,
                    removed: ddRemoved ? ddRemoved.dataItems().length : 0
                };
            """)
            
            success = state_counts and state_counts['applied'] > 0 and state_counts['removed'] > 0
            if success:
                logging.info(f"✅ Method cascade: {state_counts['applied']} applied, {state_counts['removed']} removed states")
            else:
                logging.error(f"❌ Method cascade failed")
                
            return success
            
        except Exception as e:
            logging.error(f"❌ Method cascade failed: {e}")
            return False

    def get_current_method_value(self) -> int:
        """Get current method value."""
        try:
            value = self.execute_script("""
                var ddMethod = $('#OverrideMethodId').data('kendoDropDownList');
                return ddMethod ? parseInt(ddMethod.value()) : 0;
            """)
            return value or 0
        except Exception:
            return 0

    def auto_select_first_options(self):
        """Auto-select first available options."""
        try:
            # Auto-select first method
            method_value = self.execute_script("""
                var dd = $('#OverrideMethodId').data('kendoDropDownList');
                if (dd && dd.dataItems().length > 0) {
                    var item = dd.dataItems()[0];
                    dd.value(item.Value || item.Id);
                    return item.Value || item.Id;
                }
                return null;
            """)
            
            if method_value:
                logging.info("✅ Auto-selected method")
            
            # Auto-select first states
            self.execute_script("""
                var ddApplied = $('#OverrideAppliedStateId').data('kendoDropDownList');
                if (ddApplied && ddApplied.dataItems().length > 0) {
                    var item = ddApplied.dataItems()[0];
                    ddApplied.value(item.Value || item.Id);
                }
                
                var ddRemoved = $('#OverrideRemovedStateId').data('kendoDropDownList');
                if (ddRemoved && ddRemoved.dataItems().length > 0) {
                    var item = ddRemoved.dataItems()[0];
                    ddRemoved.value(item.Value || item.Id);
                }
            """)
            
            logging.info("✅ Auto-selected states")
            return True
        except Exception as e:
            logging.error(f"❌ Auto-selection failed: {e}")
            return False

    def navigate_to_edit_overrides(self):
        """Navigate to SOC Edit Overrides page."""
        url = self.SOC_base_link + (self.SOC_id or "")
        self.driver.get(url)
        
        try:
            WebDriverWait(self.driver, self.MAX_WAIT_PAGE_LOAD_DELAY_SECONDS).until(
                EC.presence_of_element_located((By.ID, "OverrideTypeId"))
            )
            time.sleep(2)
            
            for dropdown_id in ["OverrideTypeId", "OverrideMethodId", "OverrideAppliedStateId", "OverrideRemovedStateId"]:
                self.wait_for_kendo_dropdown(dropdown_id, 10)
            
        except TimeoutException:
            logging.error("❌ Page load timeout")

    def clear_form(self):
        """Clear form fields."""
        fields = ["TagNumber", "Description", "Comment", "AdditionalValueAppliedState", "AdditionalValueRemovedState"]
        for field in fields:
            try:
                element = self.driver.find_element(By.ID, field)
                element.clear()
            except Exception:
                pass

    def fill_text_field(self, field_id: str, value: str):
        """Fill text field with value."""
        if not value:
            return
            
        try:
            element = self.driver.find_element(By.ID, field_id)
            if element.is_displayed() and element.is_enabled():
                element.clear()
                element.send_keys(str(value))
        except Exception:
            logging.warning(f"⚠️ Could not fill {field_id}")

    def map_override_type_to_value(self, override_type_text: str) -> int:
        """Map override type text to numeric value."""
        type_mapping = {
            "Программная форсировка": 2,
            "Программная": 2,
            "форсировка": 2,
            "Программная блокировка": 3,
            "блокировка": 3,
            "Аппаратная блокировка": 4,
            "Аппаратная": 4,
            "2": 2,
            "3": 3,
            "4": 4
        }
        
        if override_type_text in type_mapping:
            return type_mapping[override_type_text]
        
        for key, value in type_mapping.items():
            if key in str(override_type_text):
                return value
        
        return 2

    def add_override(self, override: dict) -> None:
        """Add a single override."""
        try:
            self.clear_form()

            # Set basic fields
            self.fill_text_field("TagNumber", override["TagNumber"])
            self.fill_text_field("Description", override["Description"])

            # Handle cascade
            if override["OverrideType"]:
                type_value = self.map_override_type_to_value(override["OverrideType"])
                
                if not self.cascade_type_to_method(type_value):
                    raise Exception("Type cascade failed")

            # Auto-select options
            self.auto_select_first_options()

            # Set optional fields
            if override["Comment"]:
                self.fill_text_field("Comment", override["Comment"])

            if override["AdditionalValueAppliedState"]:
                self.fill_text_field("AdditionalValueAppliedState", override["AdditionalValueAppliedState"])

            if override["AdditionalValueRemovedState"]:
                self.fill_text_field("AdditionalValueRemovedState", override["AdditionalValueRemovedState"])

            # Save
            add_button = self.driver.find_element(By.ID, "AddOverrideBtn")
            add_button.click()
            
            time.sleep(1)
            
            if self.check_for_errors():
                raise Exception("Validation error")
                
            logging.info(f"✅ Added: {override['TagNumber']}")

        except Exception as e:
            logging.error(f"❌ Failed: {override.get('TagNumber', 'Unknown')}: {e}")
            raise

    def check_for_errors(self) -> bool:
        """Check for error messages."""
        try:
            error_elements = self.driver.find_elements(By.CSS_SELECTOR, ".field-validation-error, .validation-summary-errors")
            for element in error_elements:
                if element.text and element.text.strip():
                    logging.warning(f"⚠️ Error: {element.text.strip()}")
                    return True
            return False
        except Exception:
            return False

    def process_all_overrides(self):
        """Process all loaded overrides."""
        total_count = len(self.list_of_overrides)
        logging.info(f"📋 Processing {total_count} overrides")
        
        success_count = 0
        failed_count = 0
        
        for index, override in enumerate(self.list_of_overrides, 1):
            try:
                logging.info(f"🔄 {index}/{total_count}: {override['TagNumber']}")
                self.add_override(override)
                success_count += 1
                
                if index < total_count:
                    time.sleep(0.5)
                
            except Exception as e:
                failed_count += 1
                logging.error(f"❌ Failed: {override['TagNumber']}")
                continue
        
        logging.info(f"🎉 Completed: {success_count} successful, {failed_count} failed")

    def wait_for_user_confirmation(self):
        msg = '⚠️  Скрипт ожидает нажатия кнопки "Подтвердить".'
        xpath = "//div[@id='bottomWindowButtons']/div"
        self.inject_info_message(msg, (By.XPATH, xpath), {'color': 'lawngreen'})
        try:
            WebDriverWait(self.driver, self.MAX_WAIT_USER_INPUT_DELAY_SECONDS).until(
                EC.title_is(self.EXPECTED_HOME_PAGE_TITLE)
            )
            logging.info("🏁 Confirm pressed, home page loaded")
        except NoSuchWindowException:
            logging.error("⚠️ User closed browser")
            self.safe_exit()
        except Exception as e:
            logging.error(f"❌ Failed waiting for confirm: {e}")
            self.inject_error_message("❌ Failed waiting for confirm")        

    def run(self, standalone=False):
        """Main execution workflow."""
        try:
            if standalone:
                self.navigate_to_base()
                self.enter_credentials_and_prepare_soc_input()
                self.wait_for_soc_input_and_submit()

            self.load_overrides_from_export()
            self.navigate_to_edit_overrides()
            
            self.SOC_locked_check()
            self.access_denied_check()
            
            self.process_all_overrides()
            self.wait_for_user_confirmation()

        except Exception as e:
            logging.error(f"❌ SOC_Importer failed: {e}")
            raise
        finally:
            if hasattr(self, 'driver'):
                self.driver.quit()


if __name__ == "__main__":
    auto_soc = SOC_Importer()
    auto_soc.run(standalone=True)