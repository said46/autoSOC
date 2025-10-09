from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import (
    NoSuchElementException, TimeoutException, ElementNotInteractableException,
    NoSuchWindowException, StaleElementReferenceException
)
import time
import openpyxl as xl

import logging
from typing_extensions import override

from base_web_bot import BaseWebBot, message_box

class autoSOC(BaseWebBot):
    def __init__(self):
        super().__init__()
        self.base_link = r"http://eptw-training.sakhalinenergy.ru"
        self.SOC_base_link = self.base_link + r"/SOC/EditOverrides/"
        self.EXPECTED_HOME_PAGE_TITLE = "СНД - Домашняя страница"    

    @override
    def load_configuration(self) -> None:
        """Load configuration - overrides parent method"""
        try:
            self.load_config_from_excel()
            # Verify the configuration was loaded
            if not hasattr(self, 'list_of_overrides') or not self.list_of_overrides:
                raise Exception("Configuration loaded but list_of_overrides is empty")
            logging.info(f"✅ Configuration verified: {len(self.list_of_overrides)} overrides ready")
        except Exception as e:
            logging.error(f"❌ Failed to load configuration: {e}")
            raise

    def load_config_from_excel(self):
        """Load configuration from Excel file"""
        try:
            logging.info("📂 Loading Excel configuration...")
            wb = xl.load_workbook('overrides.xlsx')
            
            # Load settings
            sheet = wb['Settings']
            self.user_name = sheet.cell(1, 2).value
            self.password = sheet.cell(2, 2).value
            self.time_delay = float(sheet.cell(4, 2).value)
            logging.info(f"⚙️ Settings loaded - User: {self.user_name}, Delay: {self.time_delay}")
            
            # Load overrides
            sheet = wb['overrides']
            self.list_of_overrides = []
            
            logging.info(f"📊 Excel sheet 'overrides' has {sheet.max_row} rows")
            
            for row in range(2, sheet.max_row + 1):
                tag_number = sheet.cell(row, 1).value
                if tag_number in (None, ""):
                    logging.info(f"🛑 Empty tag number at row {row}, stopping override loading")
                    break
                    
                xlsx_override = {
                    "TagNumber": tag_number,
                    "Description": sheet.cell(row, 2).value,
                    "OverrideType": sheet.cell(row, 4).value,
                    "OverrideMethod": sheet.cell(row, 5).value,
                    "Comment": sheet.cell(row, 3).value,
                    "AppliedState": sheet.cell(row, 6).value,
                    "AdditionalValueAppliedState": sheet.cell(row, 7).value,
                    "RemovedState": sheet.cell(row, 8).value,
                    "AdditionalValueRemovedState": sheet.cell(row, 9).value
                }
                self.list_of_overrides.append(xlsx_override)
            
            # Load SOC ID
            self.SOC_id = str(sheet.cell(1, 12).value)
            logging.info(f"🔢 SOC ID: {self.SOC_id}")
            
            logging.info(f"✅ Configuration loaded successfully from Excel, {len(self.list_of_overrides)} overrides to add")
            
            # Debug: Print first override to verify structure
            if self.list_of_overrides:
                logging.info(f"🔍 First override sample: {self.list_of_overrides[0]}")
            else:
                logging.warning("⚠️ No overrides were loaded from Excel file!")
                
        except Exception as e:
            logging.error(f"❌ Error loading configuration from Excel: {e}")
            self.inject_error_message("❌ Failed to load configuration from Excel")
            raise
      
    def is_menu_item_already_selected(self, parent_id, menu_item_text):
        """Check if menu item is already selected"""
        item_xpath = (
            f"//ul[@id='{parent_id}']/li[text()='{menu_item_text}' and "
            f"contains(@class ,'k-item') and contains(@class ,'k-state-selected')]"
        )
        try:
            self.driver.find_element(By.XPATH, item_xpath)
            logging.info(f"is_menu_item_already_selected: item_xpath for '{menu_item_text}', '{parent_id}' is: '{item_xpath}'")
            return True
        except NoSuchElementException:
            return False
    
    def select_menu_item(self, parent_id, menu_item_text):
        """Select an item from a menu"""
        try:
            item_xpath = f"//ul[@id='{parent_id}']/li[text()='{menu_item_text}' and contains(@class ,'k-item')]"
            logging.info(f"select_menu_item: item_xpath for '{menu_item_text}', '{parent_id}' is: '{item_xpath}'")        
            
            ignored_exceptions = (NoSuchElementException, StaleElementReferenceException)
            element = WebDriverWait(
                self.driver, 5, ignored_exceptions=ignored_exceptions
            ).until(EC.element_to_be_clickable((By.XPATH, item_xpath)))

            time.sleep(self.time_delay)
            self.driver.execute_script("arguments[0].click();", element)
            
        except NoSuchElementException:
            logging.info(f"select_menu_item: NoSuchElementException, XPATH = '{item_xpath}'")
            self.inject_error_message('NoSuchElementException: ' + item_xpath)
            self.safe_exit()
        except TimeoutException as e:
            exception_name = type(e).__name__
            logging.info(f"select_menu_item: {exception_name}, XPATH = '{item_xpath}'")
            self.inject_error_message('TimeoutException: ' + item_xpath)
            self.safe_exit()
        except ElementNotInteractableException as e:
            exception_name = type(e).__name__
            logging.info(f"select_menu_item: {exception_name}, XPATH = '{item_xpath}'")
            self.inject_error_message(f"{exception_name}: {item_xpath}")
            self.safe_exit()
        except StaleElementReferenceException as e:
            exception_name = type(e).__name__
            logging.info(f"select_menu_item: {exception_name}, XPATH = '{item_xpath}'")
            self.inject_error_message(                
                f"Исключение {exception_name}, можно нажать Confirm, чтобы сохранить те точки, "
                "которые уже добавлены, и запустить скрипт снова (предвариельно удалив уже "
                "добавленные точки из overrides.xslx)"
            )
            self.safe_exit()
    
    def login(self):
        """Perform login to the application"""
        try:
            self.driver.maximize_window()
            self.driver.get(self.base_link)
            
            self.driver.find_element(By.ID, "UserName").send_keys(self.user_name)
            self.driver.find_element(By.ID, "Password").send_keys(self.password)
            self.driver.find_element(
                By.XPATH, 
                "//button[@type='submit' and @class='panel-line-btn btn-sm k-button k-primary']"
            ).click()
            
            logging.info("✅ Login completed successfully")
            
        except Exception as e:
            logging.error(f"❌ Error during login: {e}")
            raise
    
    def navigate_to_edit_overrides(self):
        """Navigate to Edit Overrides page"""
        try:
            self.driver.get(self.SOC_base_link + self.SOC_id)
            logging.info(f"👆 Navigated to Edit Overrides page for SOC {self.SOC_id}")
            
            # Check if SOC is locked
            try:
                li_locked = self.driver.find_element(By.XPATH, "//li[contains(text(), 'Locked')]")
                self.inject_error_message('❌ SOC is locked, the script will be terminated ' + li_locked.text)
                self.safe_exit()
            except NoSuchElementException:
                logging.info("✅ Success: SOC is not locked")
            
            # Check for Access Denied
            try:
                access_denied = self.driver.find_element(By.XPATH, "//h1[text()='Access Denied']")
                self.inject_error_message(
                    access_denied.text +  
                    f'Access denied, probably SOC {self.SOC_id} is archived or in improper state')
                self.safe_exit()
            except NoSuchElementException:
                logging.info("✅ Success: no access denied issue")
                
        except Exception as e:
            logging.error(f"❌ Error navigating to Edit Overrides: {e}")
            raise
    
    def add_override(self, override: dict[str, str]) -> None:
        """Add a single override to the SOC"""
        try:
            # Enter Tag Number and Description
            self.driver.find_element(By.ID, "TagNumber").send_keys(override["TagNumber"])
            self.driver.find_element(By.ID, "Description").send_keys(override["Description"])
            
            # Select Override Type
            OverrideTypeIdMenu_XPATH = '//span[@aria-owns="OverrideTypeId_listbox"]'
            self.driver.find_element(By.XPATH, OverrideTypeIdMenu_XPATH).click()
            self.select_menu_item('OverrideTypeId_listbox', override["OverrideType"])
            
            # Select Override Method if not already selected
            if not self.is_menu_item_already_selected('OverrideMethodId_listbox', override["OverrideMethod"]):
                OverrideMethodMenu_XPATH = '//span[@aria-owns="OverrideMethodId_listbox"]'
                self.driver.find_element(By.XPATH, OverrideMethodMenu_XPATH).click()
                self.select_menu_item('OverrideMethodId_listbox', override["OverrideMethod"])
            
            # Enter Comment if provided
            if override["Comment"] is not None:
                self.driver.find_element(By.ID, "Comment").send_keys(override["Comment"])
            
            # Select Applied State
            AppliedStateMenu_XPATH = '//span[@aria-owns="OverrideAppliedStateId_listbox"]'
            self.driver.find_element(By.XPATH, AppliedStateMenu_XPATH).click()
            self.select_menu_item('OverrideAppliedStateId_listbox', override['AppliedState'])
            
            # Enter Additional Value Applied State if provided
            if override["AdditionalValueAppliedState"] is not None:
                try:
                    self.driver.find_element(By.ID, "AdditionalValueAppliedState").send_keys(
                        override["AdditionalValueAppliedState"]
                    )
                except ElementNotInteractableException as e:
                    exception_name = type(e).__name__
                    logging.info(f"send_keys() for element with ID 'AdditionalValueAppliedState': {exception_name}")
                    self.safe_exit()
            
            # Select Removed State if provided and not already selected
            if override["RemovedState"] is not None:
                if not self.is_menu_item_already_selected('OverrideRemovedStateId_listbox', override["RemovedState"]):
                    RemovedStateMenu_XPATH = '//span[@aria-owns="OverrideRemovedStateId_listbox"]'
                    self.driver.find_element(By.XPATH, RemovedStateMenu_XPATH).click()
                    self.select_menu_item('OverrideRemovedStateId_listbox', override["RemovedState"])
            
            # Enter Additional Value Removed State if provided
            if override["AdditionalValueRemovedState"] is not None:
                self.driver.find_element(By.ID, "AdditionalValueRemovedState").send_keys(
                    override["AdditionalValueRemovedState"]
                )
            
            # Click Add button
            self.driver.find_element(By.ID, "AddOverrideBtn").click()
            
            logging.info(f"✅ Added override: {override['TagNumber']}")
            
        except Exception as e:
            logging.error(f"❌ Error adding override {override.get('TagNumber', 'Unknown')}: {e}")
            raise
    
    def process_all_overrides(self):
        """Process all overrides from the Excel file"""
        logging.info(f"📋 Starting to process {len(self.list_of_overrides)} overrides")
        try:
            for i, override in enumerate(self.list_of_overrides, 1):
                logging.info(f"🔄 Processing override {i}/{len(self.list_of_overrides)}: {override['TagNumber']}")           
                # Debug: log override details
                logging.debug(f"Override details: {override}")
                self.add_override(override)
                logging.info(f"✅ Successfully processed override {i}: {override['TagNumber']}")

            logging.info("✅ All overrides processed successfully")
            
        except Exception as e:
            logging.error(f"❌ Error processing overrides: {e}")
            raise

        self.wait_for_user_confirmation()
    
    def wait_for_user_confirmation(self):
        """Wait for user to press confirm button"""
        msg = '⚠️  Скрипт ожидает нажатия кнопки "Подтвердить".'
        xpath = "//div[@id='bottomWindowButtons']/div"
        self.inject_info_message(msg, (By.XPATH, xpath), {'color': 'lawngreen'})
        try:
            # after injecting the text, the script waits for MAX_WAIT_USER_INPUT_DELAY_SECONDS minutes for the web page title to be changed
            # to "СНД - Домашняя страница", to ensure the user pressed the confirm button
            WebDriverWait(self.driver, self.MAX_WAIT_USER_INPUT_DELAY_SECONDS).until(EC.title_is(self.EXPECTED_HOME_PAGE_TITLE))
            logging.info("🏁  Confirm pressed, home page loaded")
        except NoSuchWindowException as e:
            logging.error(f"⚠️  User closed the browser window.")
            self.safe_exit()
        except Exception as e:
            logging.error(f"❌ Failed to wait for user input ('Confirm' button): {str(e)}")
            self.inject_error_message(f"❌ Failed to wait for user input ('Confirm' button){self.ERROR_MESSAGE_ENDING}.")
    
    def run(self):
        """Main method to run the automation"""
        try:
            logging.info("🚀 Starting autoSOC automation")
            
            self.login()
            self.navigate_to_edit_overrides()            
            self.process_all_overrides()
                       
            logging.info("🏁 autoSOC automation completed successfully")
        except Exception as e:
            logging.error(f"❌ autoSOC automation failed: {e}")
            self.inject_error_message("Error " + f"Automation failed: {e}")
        self.driver.quit() 

# Main execution
if __name__ == "__main__":
    auto_soc = autoSOC()
    auto_soc.run()