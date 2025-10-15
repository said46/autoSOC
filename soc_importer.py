from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (TimeoutException, NoSuchElementException)
import openpyxl as xl
import logging

from base_web_bot import BaseWebBot
# Removed duplicate import: from error_types import ErrorLevel
from soc_base_mixin import SOC_BaseMixin
from error_types import ErrorLevel, OperationResult # Import both once

class SOC_Importer(SOC_BaseMixin):
    """SOC overrides importer with proper cascade handling."""

    # =========================================================================
    # INITIALIZATION & CONFIGURATION
    # =========================================================================

    def __init__(self, soc_id=None):
        BaseWebBot.__init__(self)
        SOC_BaseMixin.__init__(self)
        self._initialized = False
        
        # Load configuration - critical for operation
        success, error_msg, severity = self.load_configuration()
        if not success:
            logging.error(f"❌ Importer initialization failed: {error_msg}")
            # Can't use inject_error_message here - browser not ready yet
            print(f"❌ FATAL: {error_msg}")
            raise RuntimeError(f"Importer initialization failed: {error_msg}")
            
        self.SOC_base_link = self.base_link + r"Soc/EditOverrides/"
        if soc_id:
            self.SOC_id = soc_id        
        self.import_file_name: str = "soc_import_export/overrides.xlsx"
        self.override_records = []
        
        self._initialized = True

    @property
    def base_link(self) -> str:
        return self._base_link

    def load_configuration(self) -> OperationResult:
        """Returns (success, error_message, severity)"""
        try:
            import configparser
            config = configparser.ConfigParser(interpolation=None)
            config.read(self.config_file, encoding="utf8")

            # ✅ Common configuration (includes SOC_id)
            success, error_msg, severity = self.load_common_configuration(config)
            if not success:
                return False, error_msg, severity

            return True, None, None

        except Exception as e:
            return False, f"Configuration failed: {e}", ErrorLevel.FATAL

    # =========================================================================
    # DATA LOADING & MAPPING
    # =========================================================================

    def load_override_records(self) -> OperationResult:
        """Load override records from Excel with user-edited Russian text."""
        try:
            wb = xl.load_workbook(self.import_file_name)
            sheet = wb.active

            self.override_records = []
            start_row = 7  # Skip instruction rows
            
            for row in range(start_row, sheet.max_row + 1):
                tag_number = sheet.cell(row, 1).value
                if not tag_number:
                    continue  # Skip empty rows

                record = {
                    # User-edited fields (direct from Excel)
                    'tag_number': str(tag_number).strip(),
                    'description': str(sheet.cell(row, 2).value or '').strip(),
                    'type_text': str(sheet.cell(row, 3).value or '').strip(),
                    'method_text': str(sheet.cell(row, 4).value or '').strip(),
                    'comment': str(sheet.cell(row, 5).value or '').strip(),
                    'applied_state_text': str(sheet.cell(row, 6).value or '').strip(),
                    'applied_additional_value': str(sheet.cell(row, 7).value or '').strip(),
                    'removed_state_text': str(sheet.cell(row, 8).value or '').strip(),
                    'removed_additional_value': str(sheet.cell(row, 9).value or '').strip()
                }
                
                self.override_records.append(record)

            logging.info(f"📋 Loaded {len(self.override_records)} user-edited override records")
            
            # Validate that mapping will work
            success, error_msg = self._validate_user_inputs()
            if not success:
                return False, error_msg, ErrorLevel.RECOVERABLE
                
            return True, None, None
            
        except Exception as e:
            error_msg = f"Failed to load override records from {self.import_file_name}: {e}"
            logging.error(error_msg)
            return False, error_msg, ErrorLevel.FATAL

    def _validate_user_inputs(self) -> tuple[bool, str]:
        """Validate that user-edited Russian text can be properly mapped."""
        valid_types = ['байпас', 'блокировка', 'форсировка', 'логики', 'сигнализации']
        
        for record in self.override_records:
            type_text = record['type_text'].lower()
            
            # Check if type text is mappable
            if type_text and not any(valid_type in type_text for valid_type in valid_types):
                logging.warning(f"⚠️ Unrecognized override type: '{record['type_text']}'")
                # Continue anyway - maybe user knows what they're doing
        
        return True, None

    def map_override_type(self, type_text: str) -> int:
            """Using generator expression for memory efficiency."""
            if not type_text:
                return None
                
            mapping_rules = (
                ("байпас", 1),
                ("блокировка", 2),
                ("форсировка", 3),
                ("логики", 4),
                ("сигнализации", 5)
            )
            
            text_lower = type_text.lower()
            
            # Generator expression - stops at first match
            match = next((value for key, value in mapping_rules if key in text_lower), None)
            return match        

    # =========================================================================
    # WEB ELEMENT INTERACTION METHODS
    # =========================================================================

    def fill_text_field(self, field_id: str, value: str) -> OperationResult:
        """Fill text field with value. Returns (success, error_message, severity)"""
        # Check browser before operation
        if not self.is_browser_alive():
            return False, "Browser closed", ErrorLevel.TERMINAL
            
        try:
            element = self.driver.find_element(By.ID, field_id)
            element.clear()
            element.send_keys(str(value))
            logging.info(f"✅ Text field filled: {field_id} = {value}")
            return True, None, None
        except NoSuchElementException:
            error_msg = f"Text field not found: {field_id}"
            logging.warning(error_msg)
            return False, error_msg, ErrorLevel.FATAL
        except Exception as e:
            error_msg = f"Failed to fill text field {field_id}: {e}"
            logging.error(error_msg)
            return False, error_msg, ErrorLevel.FATAL

    # =========================================================================
    # DROPDOWN MANAGEMENT METHODS
    # =========================================================================

    def get_dropdown_data(self, dropdown_id: str) -> dict:
        """Get complete dropdown data including items and selection."""
        # Check browser before operation
        if not self.is_browser_alive():
            return False, "Browser closed", ErrorLevel.TERMINAL
            
        try:
            data = self.driver.execute_script(f"""
                var dd = $('#{dropdown_id}').data('kendoDropDownList');
                if (!dd) return {{error: 'not_initialized'}};
                
                var items = dd.dataItems();
                var selectedValue = dd.value();
                var selectedText = dd.text();
                
                return {{
                    items: items.map(item => ({{
                        text: item.Text || item.Title,
                        value: item.Value || item.Id
                    }})),
                    selected_value: selectedValue,
                    selected_text: selectedText,
                    item_count: items.length
                }};
            """)
            return data
        except Exception as e:
            logging.error(f"❌ Failed to get dropdown data for {dropdown_id}: {e}")
            return {'error': str(e)}

    def find_dropdown_item_by_text(self, dropdown_id: str, search_text: str) -> dict:
        """Find dropdown item by partial text match."""
        # Check browser before operation
        if not self.is_browser_alive():
            return False, "Browser closed", ErrorLevel.TERMINAL
            
        data = self.get_dropdown_data(dropdown_id)
        if data.get('error') or not data.get('items'):
            return None
            
        search_lower = search_text.lower()
        for item in data['items']:
            if search_lower in item['text'].lower():
                return item
        return None

    def set_dropdown_value(self, dropdown_id: str, value: int) -> bool:
        """Set dropdown value without triggering cascade."""
        # Check browser before operation
        if not self.is_browser_alive():
            return False, "Browser closed", ErrorLevel.TERMINAL
            
        try:
            self.driver.execute_script(f"""
                var dd = $('#{dropdown_id}').data('kendoDropDownList');
                if (dd) {{
                    dd.value({value});
                }}
            """)
            logging.info(f"✅ Dropdown set: {dropdown_id} = {value}")
            return True
        except Exception as e:
            logging.error(f"❌ Failed to set {dropdown_id}: {e}")
            return False

    def trigger_cascade_change(self, dropdown_id: str, new_value: int) -> bool:
        """Trigger cascade by simulating dropdown change event."""
        # Check browser before operation
        if not self.is_browser_alive():
            return False, "Browser closed", ErrorLevel.TERMINAL
            
        try:
            # Check browser state again before proceeding with cascade logic
            if not self.is_browser_alive():
                return False, "Browser closed", ErrorLevel.TERMINAL
                
            success = self.driver.execute_script(f"""
                var source = $('#{dropdown_id}').data('kendoDropDownList');
                var element = $('#{dropdown_id}')[0];
                
                if (source && element) {{
                    var events = $._data(element, 'events');
                    if (events && events.change && events.change[0]) {{
                        var handler = events.change[0].handler;
                        source.value({new_value});
                        
                        var event = $.Event('change');
                        Object.assign(event, {{
                            target: element,
                            currentTarget: element,
                            sender: source,
                            value: {new_value}
                        }});
                        
                        handler.call(element, event);
                        return true;
                    }}
                }}
                return false;
            """)
                        
            if success:
                logging.info(f"🔄 Cascade triggered: {dropdown_id} → {new_value}")
            
            # Check again if browser is still open before waiting
            if not self.is_browser_alive():
                return False, "Browser closed", ErrorLevel.TERMINAL
                
            # Wait for common dependent widgets to be ready
            dependent_widgets = ['OverrideMethodId', 'OverrideAppliedStateId', 'OverrideRemovedStateId']
            self.wait_for_page_fully_ready(
                check_dom=False,  # DOM is already ready
                check_jquery=False,  # jQuery is already loaded
                check_kendo=False,  # Kendo is already loaded
                specific_widgets=dependent_widgets,
                timeout=10  # Shorter timeout for cascade
            )

            return success
            
        except Exception as e:
            logging.error(f"❌ Cascade failed for {dropdown_id}: {e}")
            return False

    # =========================================================================
    # FORM FILLING & VALIDATION METHODS
    # =========================================================================

    def handle_dynamic_additional_fields(self, record: dict) -> OperationResult:
        """Handle additional value fields that have dynamic visibility. Returns (success, error_message, severity)"""
        # Check browser before operation
        if not self.is_browser_alive():
            return False, "Browser closed", ErrorLevel.TERMINAL
            
        try:
            # Applied additional value
            if record['applied_additional_value']:
                if self._wait_for_element_visibility("AdditionalValueAppliedState", timeout=3):
                    success, error_msg, severity = self.fill_text_field("AdditionalValueAppliedState", record['applied_additional_value'])
                    if not success:
                        return False, error_msg, severity
                    logging.info(f"✅ Applied Additional Value set: {record['applied_additional_value']}")
                else:
                    logging.warning("⚠️ Applied Additional Value field not visible - may need different state selection")
            
            # Removed additional value  
            if record['removed_additional_value']:
                if self._wait_for_element_visibility("AdditionalValueRemovedState", timeout=3):
                    success, error_msg, severity = self.fill_text_field("AdditionalValueRemovedState", record['removed_additional_value'])
                    if not success:
                        return False, error_msg, severity
                    logging.info(f"✅ Removed Additional Value set: {record['removed_additional_value']}")
                else:
                    logging.warning("⚠️ Removed Additional Value field not visible - may need different state selection")
            
            return True, None, None
            
        except Exception as e:
            error_msg = f"Failed to handle dynamic additional fields: {e}"
            logging.error(error_msg)
            return False, error_msg, ErrorLevel.FATAL

    def fill_basic_fields(self, record: dict) -> OperationResult:
        """Fill basic text fields in the override form. Returns (success, error_message, severity)"""
        # Check browser before operation
        if not self.is_browser_alive():
            return False, "Browser closed", ErrorLevel.TERMINAL
            
        try:
            logging.info(f"📝 Filling basic fields for: {record['tag_number']}")
            
            # Fill Tag Number (required field)
            if record['tag_number']:
                success, error_msg, severity = self.fill_text_field("TagNumber", record['tag_number'])
                if not success:
                    return False, error_msg, severity
            
            # Fill Description (optional field)
            if record['description']:
                success, error_msg, severity = self.fill_text_field("Description", record['description'])
                if not success:
                    return False, error_msg, severity
            
            # Fill Comment (optional field)
            if record['comment']:
                success, error_msg, severity = self.fill_text_field("Comment", record['comment'])
                if not success:
                    return False, error_msg, severity
            
            logging.info(f"✅ Basic fields filled for: {record['tag_number']}")
            return True, None, None
            
        except Exception as e:
            error_msg = f"Failed to fill basic fields for {record['tag_number']}: {e}"
            logging.error(error_msg)
            return False, error_msg, ErrorLevel.FATAL    

    def fill_optional_fields(self, record: dict) -> OperationResult:
        """Fill optional fields in the override form. Returns (success, error_message, severity)"""
        # Check browser before operation
        if not self.is_browser_alive():
            return False, "Browser closed", ErrorLevel.TERMINAL
            
        try:
            # Fill Removed State if provided
            if record['removed_state_text']:
                removed_item = self.find_dropdown_item_by_text("OverrideRemovedStateId", record['removed_state_text'])
                if removed_item:
                    if not self.set_dropdown_value("OverrideRemovedStateId", removed_item['value']):
                        logging.warning(f"⚠️ Failed to set Removed State for {record['tag_number']}")
                else:
                    logging.warning(f"⚠️ Removed state not found: {record['removed_state_text']}")
            
            logging.info(f"✅ Optional fields processed for: {record['tag_number']}")
            return True, None, None
            
        except Exception as e:
            error_msg = f"Failed to fill optional fields for {record['tag_number']}: {e}"
            logging.error(error_msg)
            return False, error_msg, ErrorLevel.FATAL            
    
    def fill_override_form(self, record: dict) -> OperationResult:
        """Fill the override form with dynamic field visibility handling."""
        # Check browser before operation
        if not self.is_browser_alive():
            return False, "Browser closed", ErrorLevel.TERMINAL
            
        try:
            logging.info(f"🔧 Processing: {record['tag_number']}")
            
            # Step 1: Fill basic fields
            success, error_msg, severity = self.fill_basic_fields(record)
            if not success:
                return False, error_msg, severity
            
            # Step 2-4: Handle cascading dropdowns
            success, error_msg, severity = self.handle_cascading_dropdowns(record)
            if not success:
                return False, error_msg, severity
            
            # Step 5: Fill optional fields
            success, error_msg, severity = self.fill_optional_fields(record)
            if not success:
                return False, error_msg, severity
                
            # Step 6: Handle dynamic additional fields
            success, error_msg, severity = self.handle_dynamic_additional_fields(record)
            if not success:
                return False, error_msg, severity
            
            return True, None, None
            
        except Exception as e:
            error_msg = f"Form fill failed for {record['tag_number']}: {e}"
            logging.error(error_msg)
            return False, error_msg, ErrorLevel.FATAL

    def handle_cascading_dropdowns(self, record: dict) -> OperationResult:
        """Handle the Type → Method → Applied State cascade."""
        # Check browser before operation
        if not self.is_browser_alive():
            return False, "Browser closed", ErrorLevel.TERMINAL
        
        # Step 2: Handle Type → Method cascade
        if record['type_text']:
            success, error_msg, severity = self.process_type_method_cascade(record)
            if not success:
                return False, error_msg, severity
        
        # Step 4: Handle Applied State selection
        if record['applied_state_text']:
            success, error_msg, severity = self.process_applied_state(record)
            if not success:
                return False, error_msg, severity
        
        return True, None, None

    def process_type_method_cascade(self, record: dict) -> OperationResult:
        """Process Type → Method cascade with smart waiting."""
        # Check browser before operation
        if not self.is_browser_alive():
            return False, "Browser closed", ErrorLevel.TERMINAL
            
        type_value = self.map_override_type(record['type_text'])
        if not type_value:
            return True, None, None  # No type mapping found is acceptable
        
        if not self.trigger_cascade_change("OverrideTypeId", type_value):
            return False, "Failed to trigger Type cascade", ErrorLevel.FATAL
        
        # Only wait if we actually need to select a method
        if record['method_text']:
            # Wait for Method dropdown specifically since we need to use it
            success = self.wait_for_page_fully_ready(
                check_dom=False,
                check_jquery=False,
                check_kendo=False,
                specific_widgets=['OverrideMethodId'],
                timeout=10
            )
            
            if not success:
                logging.warning("⚠️ Method dropdown not ready, but continuing...")
        
        # Step 3: Select specific Method (if needed)
        if record['method_text']:
            return self.process_method_selection(record)
        
        return True, None, None

    def process_method_selection(self, record: dict) -> OperationResult:
        """Process method selection and trigger cascade."""
        # Check browser before operation
        if not self.is_browser_alive():
            return False, "Browser closed", ErrorLevel.TERMINAL
            
        method_item = self.find_dropdown_item_by_text("OverrideMethodId", record['method_text'])
        if not method_item:
            return False, f"Method not found: {record['method_text']}", ErrorLevel.FATAL
        
        if not self.trigger_cascade_change("OverrideMethodId", method_item['value']):
            return False, "Failed to trigger Method cascade", ErrorLevel.FATAL
        
        # Wait for dependent state widgets to update instead of sleeping
        dependent_widgets = ['OverrideAppliedStateId', 'OverrideRemovedStateId']
        success = self.wait_for_page_fully_ready(
            check_dom=False,  # DOM is already ready
            check_jquery=False,  # jQuery is already loaded
            check_kendo=False,  # Kendo is already loaded
            specific_widgets=dependent_widgets,
            timeout=10  # Shorter timeout for cascade
        )
        
        if not success:
            logging.warning("⚠️ State widgets not fully ready after method cascade (may be normal for some methods)")
        
        return True, None, None

    def process_applied_state(self, record: dict) -> OperationResult:
        """Process applied state selection."""
        # Check browser before operation
        if not self.is_browser_alive():
            return False, "Browser closed", ErrorLevel.TERMINAL
            
        applied_item = self.find_dropdown_item_by_text("OverrideAppliedStateId", record['applied_state_text'])
        if not applied_item:
            return False, f"Applied state not found: {record['applied_state_text']}", ErrorLevel.FATAL
        
        if not self.set_dropdown_value("OverrideAppliedStateId", applied_item['value']):
            return False, "Failed to set Applied State", ErrorLevel.FATAL
        
        if not self.trigger_cascade_change("OverrideAppliedStateId", applied_item['value']):
            return False, "Failed to trigger Applied State cascade", ErrorLevel.FATAL
        
        return True, None, None

    def submit_override(self, record: dict) -> OperationResult:
        """Submit the completed override form. Returns (success, error_message, severity)"""
        # Check browser before operation
        if not self.is_browser_alive():
            return False, "Browser closed", ErrorLevel.TERMINAL
            
        try:           
            if not self.click_button((By.ID, "AddOverrideBtn")):
                error_msg = "Add Override button not found or not clickable"
                logging.error(f"❌ {error_msg}")
                return False, error_msg, ErrorLevel.FATAL            
            return True, None, None
        except Exception as e:
            error_msg = f"Submit override failed for {record['tag_number']}: {e}"
            logging.error(error_msg)
            return False, error_msg, ErrorLevel.FATAL

    # =========================================================================
    # WORKFLOW EXECUTION METHODS
    # =========================================================================

    def process_single_override(self, record: dict) -> OperationResult:
        """Process a single override record end-to-end. Returns (success, error_message, severity)"""
        # Check browser before operation - critical for loop operations
        if not self.is_browser_alive():
            return False, "Browser closed", ErrorLevel.TERMINAL
                
        try:
            logging.info(f"🔍 START process_single_override for: {record['tag_number']}")
            
            # Fill the form
            logging.info(f"🔍 Calling fill_override_form for: {record['tag_number']}")
            success, error_msg, severity = self.fill_override_form(record)
            if not success:
                logging.error(f"❌ fill_override_form returned False for: {record['tag_number']}")
                return False, error_msg, severity
            
            logging.info(f"🔍 fill_override_form completed, calling submit_override for: {record['tag_number']}")
            
            # Submit the override
            success, error_msg, severity = self.submit_override(record)
            if not success:
                logging.error(f"❌ submit_override returned False for: {record['tag_number']}")
                return False, error_msg, severity
            
            logging.info(f"✅ process_single_override completed successfully for: {record['tag_number']}")
            return True, None, None
            
        except Exception as e:
            error_msg = f"Uncaught exception in process_single_override for {record['tag_number']}: {e}"
            logging.error(error_msg)
            return False, error_msg, ErrorLevel.FATAL

    def execute_import_workflow(self) -> OperationResult:
        """Main import workflow execution. Returns (success, error_message, severity)"""
        try:
            total_count = len(self.override_records)
            success_count = 0
            
            logging.info(f"🚀 Starting import of {total_count} overrides")
            
            for i, record in enumerate(self.override_records, 1):
                # Check browser BEFORE each iteration - THIS IS CRITICAL
                if not self.is_browser_alive():
                    return False, "Browser closed", ErrorLevel.TERMINAL
                    
                logging.info(f"📝 Processing {i}/{total_count}: {record['tag_number']}")
                
                success, error_msg, severity = self.process_single_override(record)
                if success:
                    success_count += 1
                else:
                    logging.error(f"❌ Failed: {record['tag_number']}: {error_msg}")
                    # Inject error message
                    self.inject_info_message(f"Failed: {record['tag_number']}")
            
            logging.info(f"🎉 Import completed: {success_count}/{total_count} successful")
            return True, None, None
            
        except Exception as e:
            error_msg = f"Import workflow failed: {e}"
            logging.error(error_msg)
            return False, error_msg, ErrorLevel.FATAL

    # =========================================================================
    # NAVIGATION & WAIT METHODS
    # =========================================================================

    def navigate_to_edit_overrides(self) -> OperationResult:
        """Navigate to SOC Edit Overrides page with complete error handling."""
        # Check browser before operation
        if not self.is_browser_alive():
            return False, "Browser closed", ErrorLevel.TERMINAL
            
        try:
            url = self.SOC_base_link + self.SOC_id
            logging.info(f"🌐 Navigating to: {url}")
            
            self.driver.get(url)
            
            if not self.wait_for_page_fully_ready(check_jquery=False, check_kendo=False):
                return False, "Edit overrides page failed to load", ErrorLevel.FATAL

            # Security and access checks
            success, error_msg = self.error_404_not_present_check()
            if not success:
                return False, error_msg, ErrorLevel.FATAL
            
            success, error_msg = self.SOC_locked_check()
            if not success:
                return False, error_msg, ErrorLevel.FATAL
                       
            success, error_msg = self.access_denied_check()
            if not success:
                return False, error_msg, ErrorLevel.FATAL            
                        
            if not self._verify_edit_overrides_page():
                return False, "Edit overrides page verification failed", ErrorLevel.FATAL
            
            logging.info("✅ Successfully navigated to SOC Edit Overrides")
            return True, None, None
            
        except TimeoutException:
            error_msg = "Timeout waiting for edit overrides page to load"
            logging.error(error_msg)
            return False, error_msg, ErrorLevel.FATAL                
        except Exception as e:
            error_msg = f"Navigation to edit overrides failed: {e}"
            logging.error(error_msg)
            return False, error_msg, ErrorLevel.FATAL

    def _verify_edit_overrides_page(self) -> bool:
        """Verify we're actually on the Edit Overrides page."""
        # Check browser before operation
        if not self.is_browser_alive():
            return False, "Browser closed", ErrorLevel.TERMINAL
            
        try:
            current_url = self.driver.current_url
            if "/Soc/EditOverrides/" not in current_url:
                logging.error(f"❌ Wrong page loaded: {current_url}")
                return False
                                           
            logging.info("✅ Verified Edit Overrides page content")
            return True            
        except Exception as e:
            logging.error(f"❌ Page verification failed: {e}")
            return False

    def wait_for_user_confirmation(self) -> OperationResult:
        """Wait for user confirmation. Returns (success, error_message, severity)"""
        # Check browser before operation
        if not self.is_browser_alive():
            return True, "Browser already closed by user", ErrorLevel.RECOVERABLE
            
        try:
            msg = '⚠️ Скрипт ожидает нажатия кнопки "Подтвердить".'
            xpath = "//div[@id='bottomWindowButtons']/div"
            
            success = self._inject_message(msg, (By.XPATH, xpath), {'color': 'lawngreen'})
            if not success:
                logging.warning("⚠️ Failed to inject confirmation message, but continuing...")
            
            WebDriverWait(self.driver, self.MAX_WAIT_USER_INPUT_DELAY_SECONDS).until(
                EC.title_is(self.EXPECTED_HOME_PAGE_TITLE)
            )
            logging.info("🏁 Confirm pressed, home page loaded")
            return True, None, None
            
        except TimeoutException:
            # Timeout is normal - user didn't press confirm within the time limit
            if not self.is_browser_alive():
                return True, "Browser closed by user during confirmation wait", ErrorLevel.RECOVERABLE
            return True, "User confirmation timeout - continuing anyway", ErrorLevel.RECOVERABLE
            
        except Exception as e:
            # If browser closed during wait, that's normal user behavior
            if not self.is_browser_alive():
                return True, "Browser closed by user during confirmation wait", ErrorLevel.RECOVERABLE
            error_msg = f"Failed waiting for confirm: {e}"
            logging.error(error_msg)
            return False, error_msg, ErrorLevel.FATAL

    def run(self, standalone=False):
        """Main execution workflow."""
        if not self._initialized:
            logging.error("❌ Importer not properly initialized")
            return
            
        try:
            if standalone:
                self.navigate_to_base()
                self.enter_credentials_and_prepare_soc_input()
                
                success, error_msg = self.wait_for_soc_input_and_submit()
                if not success:
                    if not self._handle_result(False, error_msg, ErrorLevel.FATAL):
                        return                

            # Main workflow with proper severity handling
            success, error_msg, severity = self.load_override_records()
            if not self._handle_result(success, error_msg, severity):
                return
                
            success, error_msg, severity = self.navigate_to_edit_overrides()
            if not self._handle_result(success, error_msg, severity):
                return

            success, error_msg, severity = self.execute_import_workflow()
            if not self._handle_result(success, error_msg, severity):
                return

            if not self.is_browser_alive():
                    logging.info("🏁 Browser closed by user - import completed successfully")
                    return            

            self.inject_info_message("Import finished, check the message below new confirm button ", style_addons={'color': 'darkorange'})
            
            # Wait for user to press "Confirm" - browser stays open after this
            success, error_msg, severity = self.wait_for_user_confirmation()
            if not self._handle_result(success, error_msg, severity):
                return
                
            # No safe_exit() - let user decide when to close browser
            logging.info("🏁 Import workflow completed - browser remains open for user")
            
        except Exception as e:
            logging.error(f"❌ Unhandled exception in main workflow: {e}")
            if self.is_browser_alive():
                self.safe_exit()
            
if __name__ == "__main__":
    try:
        bot = SOC_Importer()
        bot.run(standalone=True)
    except Exception as e:
        print(f"❌ Failed to start importer: {e}")
        logging.error(f"❌ Importer startup failed: {e}")