# soc_exporter.py
from selenium.webdriver.common.by import By
from selenium.common.exceptions import (NoSuchElementException, TimeoutException, NoSuchWindowException, WebDriverException)
from selenium.webdriver.support.wait import WebDriverWait
import logging
import configparser
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

from base_web_bot import BaseWebBot
from soc_base_mixin import SOC_BaseMixin
from error_types import ErrorLevel, OperationResult

class SOC_Exporter(SOC_BaseMixin):
    """
    Specialized bot for exporting SOC overrides table to Excel.
    Exports data in the same format expected by the SOC_Importer.
    
    Improved with better Kendo UI integration based on app.min.js insights:
    - Proper grid initialization and data source handling
    - Multiple data extraction strategies
    - Enhanced error recovery for Kendo widgets
    """

    # =========================================================================
    # INITIALIZATION AND CONFIGURATION
    # =========================================================================

    def __init__(self, soc_id=None):
        BaseWebBot.__init__(self)
        SOC_BaseMixin.__init__(self)
        self._initialized = False
        
        # Load configuration - critical for operation
        success, error_msg, severity = self.load_configuration()
        if not success:
            logging.error(f"❌ Exporter initialization failed: {error_msg}")
            # Can't use inject_error_message here - browser not ready yet
            print(f"❌ FATAL: {error_msg}")
            raise RuntimeError(f"Exporter initialization failed: {error_msg}")
            
        self.SOC_ID_PATTERN = r"^\d{7,8}$"
        if soc_id:
            self.SOC_id = soc_id
            
        self._initialized = True

    def load_configuration(self) -> OperationResult:
        """Returns (success, error_message, severity)"""
        try:
            config = configparser.ConfigParser(interpolation=None)
            config.read(self.config_file, encoding="utf8")

            # ✅ Common configuration (includes SOC_id)
            success, error_msg, severity = self.load_common_configuration(config)
            if not success:
                return False, error_msg, severity

            logging.info(f"✅ Configuration loaded from {self.config_file}")
            return True, None, None

        except Exception as e:
            return False, f"Configuration failed: {e}", ErrorLevel.FATAL

    # =========================================================================
    # PROPERTIES AND SETTERS
    # =========================================================================

    @property
    def base_link(self) -> str:
        return self._base_link

    def set_soc_id(self, soc_id: str) -> None:
        self.SOC_id = soc_id

    # =========================================================================
    # NAVIGATION METHODS
    # =========================================================================

    def navigate_to_soc_details(self) -> OperationResult:
        """Returns (success, error_message, severity)"""
        # Check browser state before navigation
        if not self.is_browser_alive():
            return False, "Browser closed", ErrorLevel.TERMINAL
            
        try:
            soc_details_url = self.base_link + f"Soc/Details/{self.SOC_id}"
            logging.info(f"🌐 Navigating to: {soc_details_url}")

            self.driver.get(soc_details_url)

            success, error_msg = self.error_404_not_present_check()
            if not success:
                return False, error_msg, ErrorLevel.FATAL            

            # 🔥 SINGLE WAIT CALL FOR EVERYTHING
            if not self.wait_for_page_fully_ready(
                specific_widgets=['Overrides']  # Wait specifically for the grid
            ):
                return False, "Page failed to load completely", ErrorLevel.RECOVERABLE

            if not self.error_404_not_present_check():
                return False, "Edit overrides page verification failed", ErrorLevel.FATAL
            
            is_correct_page, error_msg = self.url_contains_SOC_Details_check()
            if not is_correct_page:
                return False, error_msg, ErrorLevel.FATAL

            logging.info("✅ Successfully navigated to SOC Details with all widgets ready")
            return True, None, None
            
        except TimeoutException:
            error_msg = "Timeout waiting for SOC details page to load"
            logging.error(error_msg)
            return False, error_msg, ErrorLevel.FATAL
        except Exception as e:
            error_msg = f"Navigation failed: {e}"
            logging.error(error_msg)
            return False, error_msg, ErrorLevel.FATAL

    # =========================================================================
    # DATA EXTRACTION METHODS
    # =========================================================================

    def wait_for_grid_data_loaded(self, grid_selector='#Overrides', timeout=15) -> bool:
        """
        Wait specifically for the Kendo grid to finish loading its data.
        IMPROVED: Added Kendo UI initialization checks and data refresh triggers
        
        Returns True if data is loaded, False on timeout.
        """
        try:
            wait = WebDriverWait(self.driver, timeout)
            
            def grid_data_loaded(driver):
                try:
                    script = f"""
                    var grid = $('{grid_selector}').data('kendoGrid');
                    
                    // Strategy 1: Check if grid exists and has data
                    if (grid && grid.dataSource) {{
                        var dataSource = grid.dataSource;
                        
                        // If no data but not loading, trigger refresh
                        if (dataSource.data().length === 0 && !dataSource._loading) {{
                            dataSource.read();
                            return false;
                        }}
                        
                        return dataSource.data().length > 0 && !dataSource._loading;
                    }}
                    
                    // Strategy 2: Grid element exists but Kendo widget not bound
                    var gridElement = $('{grid_selector}');
                    if (gridElement.length > 0 && !gridElement.data('kendoGrid')) {{
                        // Check if we need to trigger any initialization events
                        var parentContainers = gridElement.closest('[data-role]');
                        if (parentContainers.length > 0) {{
                            parentContainers.trigger('show');
                        }}
                        return false;
                    }}
                    
                    // Strategy 3: Check for loading indicators
                    var loadingIndicators = $('{grid_selector} .k-loading-mask, {grid_selector} .k-loading');
                    if (loadingIndicators.length > 0 && loadingIndicators.is(':visible')) {{
                        return false;
                    }}
                    
                    return false;
                    """
                    return driver.execute_script(script)
                except Exception as e:
                    logging.debug(f"Grid check error: {e}")
                    return False
            
            return wait.until(grid_data_loaded)
            
        except TimeoutException:
            logging.warning(f"⚠️ Timeout waiting for grid data to load after {timeout} seconds")
            # Try one more approach before giving up
            return self._fallback_grid_check(grid_selector)
        except Exception as e:
            logging.warning(f"⚠️ Error waiting for grid data: {e}")
            return False

    def _fallback_grid_check(self, grid_selector: str) -> bool:
        """
        Fallback method to check grid data when primary method times out.
        Uses alternative approaches to verify grid state.
        """
        try:
            # Check if grid element exists and has any visible rows
            script = f"""
            var gridElement = $('{grid_selector}');
            if (gridElement.length === 0) return false;
            
            // Check for visible rows in the grid
            var visibleRows = gridElement.find('tr[data-uid]:visible');
            if (visibleRows.length > 0) return true;
            
            // Check for "no data" message
            var noDataMessage = gridElement.find('.k-nodata, .no-data');
            if (noDataMessage.length > 0 && noDataMessage.is(':visible')) {{
                return true; // Grid is loaded but empty
            }}
            
            return false;
            """
            return self.driver.execute_script(script)
        except Exception as e:
            logging.debug(f"Fallback grid check failed: {e}")
            return False

    def check_if_overrides_exist(self) -> OperationResult:
        """Returns (success, error_message, severity)"""
        # Check browser state before data extraction
        if not self.is_browser_alive():
            return False, "Browser closed", ErrorLevel.TERMINAL
            
        try:
            overrides_section = self.driver.find_element(By.XPATH, "//label[contains(text(), 'Отчет по форсировкам')]")
            parent_panel = overrides_section.find_element(By.XPATH, "./ancestor::div[contains(@class, 'issow-panel')]")
            
            no_data_elements = parent_panel.find_elements(By.XPATH, ".//*[contains(text(), 'Нет данных') or contains(text(), 'No data')]")
            if no_data_elements:
                logging.info("ℹ️ No overrides data found")
                return True, "No overrides data found", ErrorLevel.RECOVERABLE
            return True, None, None

        except NoSuchElementException:
            logging.warning("⚠️ Overrides section not found")
            return False, "Overrides section not found", ErrorLevel.RECOVERABLE
        except Exception as e:
            logging.error(f"❌ Error checking overrides: {e}")
            return False, f"Error checking overrides: {e}", ErrorLevel.RECOVERABLE

    def extract_overrides_table_data(self) -> tuple[bool, list | None, list | None, str | None, ErrorLevel]:
        """
        Extract SOC overrides data matching the importer's expected format.
        IMPROVED: Multiple extraction strategies with better error recovery
        
        Expected column order for importer:
        TagNumber, Description, OverrideType, OverrideMethod, Comment, 
        AppliedState, AdditionalValueAppliedState, RemovedState, AdditionalValueRemovedState
        """
        # Check browser state before data extraction
        if not self.is_browser_alive():
            return False, None, None, "Browser closed", ErrorLevel.TERMINAL
            
        try:
            # ✅ Wait for grid data with improved method
            if not self.wait_for_grid_data_loaded():
                logging.warning("⚠️ Grid data may not be fully loaded, attempting extraction anyway")
            
            # Try multiple extraction strategies
            result = self._extract_data_strategy_primary()
            if not result:
                logging.info("🔄 Primary extraction failed, trying fallback strategy...")
                result = self._extract_data_strategy_fallback()
            
            if not result:
                return True, [], [], "No data found in grid", ErrorLevel.RECOVERABLE

            headers = result.get('headers', [])
            rows = result.get('rows', [])

            if headers and rows:
                logging.info(f"✅ Extracted {len(rows)} overrides with {len(headers)} columns")
                logging.info(f"📊 Columns: {', '.join(headers)}")
                return True, headers, rows, None, None

            return True, [], [], "No data extracted", ErrorLevel.RECOVERABLE

        except Exception as e:
            return False, None, None, f"Failed to extract data: {e}", ErrorLevel.FATAL

    def _extract_data_strategy_primary(self) -> dict:
        """
        Primary extraction strategy using Kendo Grid data source.
        This is the most reliable method when Kendo UI is properly initialized.
        """
        try:
            script = """
            function extractGridData() {
                var grid = $('#Overrides').data('kendoGrid');
                
                // Strategy 1: Direct data source access
                if (grid && grid.dataSource) {
                    var data = grid.dataSource.data();
                    if (data.length === 0) return null;

                    // Match the importer's expected column order (9 columns)
                    var headers = [
                        'TagNumber', 'Description', 'OverrideType', 'OverrideMethod', 'Comment',
                        'AppliedState', 'AdditionalValueAppliedState', 'RemovedState', 'AdditionalValueRemovedState'
                    ];

                    var rows = data.map(item => [
                        item.TagNumber || '',
                        item.Description || '',
                        item.OverrideType ? (item.OverrideType.Title || '') : '',
                        item.OverrideMethod ? (item.OverrideMethod.Title || '') : '',
                        item.Comment || '',
                        item.OverrideAppliedState ? (item.OverrideAppliedState.Title || '') : '',
                        item.AdditionalValueAppliedState || '',
                        item.OverrideRemovedState ? (item.OverrideRemovedState.Title || '') : '',
                        item.AdditionalValueRemovedState || ''
                    ]);

                    return {headers: headers, rows: rows};
                }
                return null;
            }
            return extractGridData();
            """
            return self.driver.execute_script(script)
        except Exception as e:
            logging.debug(f"Primary extraction failed: {e}")
            return None

    def _extract_data_strategy_fallback(self) -> dict:
        """
        Fallback extraction strategy using DOM parsing.
        Used when Kendo Grid data source is not accessible.
        """
        try:
            script = """
            function extractDomData() {
                var grid = $('#Overrides');
                if (grid.length === 0) return null;
                
                // Try to extract from visible table rows
                var rows = grid.find('tr[data-uid]');
                if (rows.length === 0) {
                    // Check for no data message
                    var noData = grid.find('.k-nodata, .no-data');
                    if (noData.length > 0) {
                        return {headers: [], rows: []}; // Empty grid
                    }
                    return null;
                }
                
                // Extract headers from thead
                var headerCells = grid.find('thead th[data-field]');
                var headers = [];
                headerCells.each(function() {
                    var field = $(this).data('field');
                    if (field) headers.push(field);
                });
                
                // If no headers found, use default importer order
                if (headers.length === 0) {
                    headers = [
                        'TagNumber', 'Description', 'OverrideType', 'OverrideMethod', 'Comment',
                        'AppliedState', 'AdditionalValueAppliedState', 'RemovedState', 'AdditionalValueRemovedState'
                    ];
                }
                
                // Extract data from rows
                var dataRows = [];
                rows.each(function() {
                    var row = $(this);
                    var rowData = [];
                    
                    // Extract cell data
                    row.find('td').each(function() {
                        var cellText = $(this).text().trim();
                        rowData.push(cellText);
                    });
                    
                    if (rowData.length > 0) {
                        dataRows.push(rowData);
                    }
                });
                
                return {headers: headers, rows: dataRows};
            }
            return extractDomData();
            """
            return self.driver.execute_script(script)
        except Exception as e:
            logging.debug(f"Fallback extraction failed: {e}")
            return None

    def _extract_data_strategy_ajax(self) -> dict:
        """
        Advanced strategy: Try to intercept or trigger AJAX data loading.
        This can be used if the grid loads data via separate API calls.
        """
        try:
            script = """
            function triggerDataRefresh() {
                var grid = $('#Overrides').data('kendoGrid');
                if (grid && grid.dataSource) {
                    // Force data refresh
                    grid.dataSource.read().then(function() {
                        console.log('Data refresh triggered');
                    });
                    return true;
                }
                return false;
            }
            return triggerDataRefresh();
            """
            # Trigger refresh and wait briefly
            self.driver.execute_script(script)
            self.driver.implicitly_wait(2)  # Brief wait for data refresh
            
            # Now try primary extraction again
            return self._extract_data_strategy_primary()
            
        except Exception as e:
            logging.debug(f"AJAX extraction failed: {e}")
            return None

    # =========================================================================
    # EXCEL EXPORT METHODS
    # =========================================================================

    def _auto_adjust_column_widths(self, sheet, headers: list, rows: list) -> None:
        """
        Auto-adjust column widths for better readability.
        """
        try:
            max_lengths = [len(str(header)) for header in headers]

            for row in rows:
                for col_index, cell_value in enumerate(row):
                    if col_index < len(max_lengths):
                        cell_length = len(str(cell_value))
                        if cell_length > max_lengths[col_index]:
                            max_lengths[col_index] = cell_length

            for col_index, max_len in enumerate(max_lengths, 1):
                adjusted_width = min(max_len + 2, 50)
                column_letter = get_column_letter(col_index)
                sheet.column_dimensions[column_letter].width = adjusted_width

        except Exception as e:
            logging.warning(f"⚠️ Could not adjust column widths: {e}")

    def create_excel_file(self, headers: list, rows: list, filename: str = None) -> OperationResult:
        """
        Create Excel file with fixed column order matching the importer's expectations.
        
        File structure (same as old version):
        - Rows 1-4: Metadata
        - Row 5: Empty (spacing)
        - Row 6: Headers (fixed order for importer)
        - Rows 7+: Data rows
        """
        try:
            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"soc_import_export/SOC_{self.SOC_id}_overrides_{timestamp}.xlsx"

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = f'SOC_{self.SOC_id}'

            # === METADATA (same as old version) ===
            metadata = [
                "SOC Overrides Export",
                f"SOC ID: {self.SOC_id}",
                f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                f"Total Overrides: {len(rows)}",
                "",  # Empty row for spacing (row 5)
                # Row 6 will contain headers
            ]
            
            for i, value in enumerate(metadata, 1):
                sheet.cell(row=i, column=1, value=value)

            # === HEADERS at row 6 (same as old version) ===
            header_row = 6
            for col_index, header in enumerate(headers, 1):
                cell = sheet.cell(row=header_row, column=col_index, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
                
                # Light gray background (same as old version)
                cell.fill = openpyxl.styles.PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            # === DATA starting from row 7 (same as old version) ===
            for row_index, row_data in enumerate(rows, header_row + 1):
                for col_index, cell_value in enumerate(row_data, 1):
                    sheet.cell(row=row_index, column=col_index, value=cell_value)

            # Adjust column widths for better readability
            self._auto_adjust_column_widths(sheet, headers, rows)

            workbook.save(filename)
            logging.info(f"💾 Saved to: {filename}")
            logging.info(f"📁 File ready for import with {len(rows)} overrides")
            return True, None, None

        except Exception as e:
            error_msg = f"Failed to create Excel: {e}"
            logging.error(f"❌ {error_msg}")
            return False, error_msg, ErrorLevel.FATAL

    def extract_and_export_overrides(self) -> OperationResult:
        """
        Extract overrides and export to Excel in importer-compatible format.
        IMPROVED: Better error handling and user feedback
        
        Returns (success, error_message, severity)
        """
        # Check browser state before export operation
        if not self.is_browser_alive():
            return False, "Browser closed", ErrorLevel.TERMINAL
            
        try:
            # First check if overrides exist at all
            success, error_msg, severity = self.check_if_overrides_exist()
            if not success and severity == ErrorLevel.RECOVERABLE:
                # Recoverable error - might still have data
                logging.warning(f"⚠️ Overrides check issue: {error_msg}, but continuing...")
            elif not success:
                return False, error_msg, severity

            # Extract the data
            success, headers, rows, error_msg, severity = self.extract_overrides_table_data()
            
            if not success:
                return False, error_msg, severity

            if not headers or not rows:
                msg = f"⚡ No overrides found for SOC {self.SOC_id}"
                logging.info(msg)
                self.inject_info_message(msg, style_addons={'color': 'orange'})
                return True, msg, ErrorLevel.RECOVERABLE
            else:
                success, error_msg, severity = self.create_excel_file(headers, rows)
                if success:
                    msg = f"✅ SOC {self.SOC_id} overrides exported successfully ({len(rows)} records)"
                    logging.info(msg)
                    self.inject_info_message(msg, style_addons={'color': 'darkorange'})
                    return True, None, None
                else:
                    msg = f"❌ Failed to save Excel for SOC {self.SOC_id}: {error_msg}"
                    self.inject_info_message(msg, style_addons={'color': 'red'})
                    return False, msg, severity

        except Exception as e:
            error_msg = f"Export error: {e}"
            logging.error(f"❌ {error_msg}")
            self.inject_info_message(f"❌ {error_msg}", style_addons={'color': 'red'})
            return False, error_msg, ErrorLevel.FATAL

    # =========================================================================
    # MAIN EXECUTION WORKFLOW
    # =========================================================================

    def run(self, standalone=False):
        """
        Main execution workflow for SOC export.
        IMPROVED: Better error handling and user feedback throughout
        """
        if not self._initialized:
            logging.error("❌ Exporter not properly initialized")
            return
            
        try:
            if standalone:
                self.navigate_to_base()
                self.enter_credentials_and_prepare_soc_input()
                
                success, error_msg = self.wait_for_soc_input_and_submit()
                if not success:
                    if not self._handle_result(False, error_msg, ErrorLevel.FATAL):
                        return                

            # Main workflow with proper severity handling
            logging.info(f"🚀 Starting SOC export workflow for SOC {self.SOC_id}")
            
            success, error_msg, severity = self.navigate_to_soc_details()
            if not self._handle_result(success, error_msg, severity):
                return
       
            success, error_msg, severity = self.extract_and_export_overrides()
            if not self._handle_result(success, error_msg, severity):
                return

            # Final success message
            if self.is_browser_alive():
                self.inject_info_message(
                    f"✅ SOC {self.SOC_id} export completed successfully!", 
                    style_addons={'color': 'green'}
                )
                
            logging.info("🏁 SOC export completed successfully")

        except Exception as e:
            logging.error(f"❌ Unhandled exception in main workflow: {e}")
            if self.is_browser_alive():
                self.inject_error_message(f"Export failed: {str(e)}")
            else:
                logging.info("🏁 Browser closed by user during export")

    def run_with_retry(self, standalone=False, max_retries=2):
        """
        Enhanced run method with retry capability for transient failures.
        Useful for handling occasional Kendo UI initialization issues.
        """
        for attempt in range(max_retries + 1):
            try:
                logging.info(f"🔄 Export attempt {attempt + 1}/{max_retries + 1}")
                self.run(standalone)
                break  # Success, exit retry loop
            except Exception as e:
                if attempt < max_retries:
                    logging.warning(f"⚠️ Attempt {attempt + 1} failed, retrying: {e}")
                    # Brief pause before retry
                    self.driver.implicitly_wait(2)
                else:
                    logging.error(f"❌ All export attempts failed: {e}")
                    if self.is_browser_alive():
                        self.inject_error_message(f"All export attempts failed: {str(e)}")

if __name__ == "__main__":
    try:
        bot = SOC_Exporter()
        # Use retry version for better reliability
        bot.run_with_retry(standalone=True)
    except Exception as e:
        print(f"❌ Failed to start exporter: {e}")
        logging.error(f"❌ Exporter startup failed: {e}")
