import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from datetime import datetime

def create_xlsx_from_json(json_file_path, output_file_path=None):
    """
    Create XLSX file from JSON data in the same format as SOC exporter
    """
    try:
        # Load JSON data
        with open(json_file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)

        if not data:
            print("❌ No data found in JSON file")
            return False

        # Create workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'SOC_Overrides'

        # Add metadata header (same as exporter format)
        sheet.cell(row=1, column=1, value="SOC Overrides Export")
        sheet.cell(row=2, column=1, value=f"SOC ID: From_JSON")
        sheet.cell(row=3, column=1, value=f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        sheet.cell(row=4, column=1, value=f"Total Overrides: {len(data)}")

        # Define headers (same as enhanced exporter)
        headers = [
            'TagNumber',
            'Description',
            'OverrideType',
            'OverrideMethod',
            'AppliedState',
            'RemovedState',
            'Comment',
            'AdditionalValueAppliedState',
            'AdditionalValueRemovedState',
            'CurrentState'
        ]

        # Write headers with bold formatting (starting from row 6)
        header_row = 6
        for col_index, header in enumerate(headers, 1):
            cell = sheet.cell(row=header_row, column=col_index, value=header)
            cell.font = Font(bold=True)

        # Write data rows (starting from row 7)
        for row_index, item in enumerate(data, header_row + 1):
            row_data = [
                # TagNumber
                item.get('TagNumber', ''),
                # Description
                item.get('Description', ''),
                # OverrideType (ShortForm preferred, fallback to Title)
                item.get('OverrideType', {}).get('ShortForm') or item.get('OverrideType', {}).get('Title', ''),
                # OverrideMethod (ShortForm preferred, fallback to Title)
                item.get('OverrideMethod', {}).get('ShortForm') or item.get('OverrideMethod', {}).get('Title', ''),
                # AppliedState
                item.get('OverrideAppliedState', {}).get('Title', ''),
                # RemovedState
                item.get('OverrideRemovedState', {}).get('Title', ''),
                # Comment
                item.get('Comment', ''),
                # AdditionalValueAppliedState
                item.get('AdditionalValueAppliedState', ''),
                # AdditionalValueRemovedState
                item.get('AdditionalValueRemovedState', ''),
                # CurrentState
                item.get('CurrentState', {}).get('Title', '')
            ]

            # Write each cell in the row
            for col_index, cell_value in enumerate(row_data, 1):
                sheet.cell(row=row_index, column=col_index, value=cell_value)

        # Auto-adjust column widths
        _auto_adjust_column_widths(sheet, headers, data)

        # Generate output filename if not provided
        if output_file_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file_path = f"soc_resources/SOC_FromJSON_overrides_{timestamp}.xlsx"

        # Save workbook
        workbook.save(output_file_path)
        print(f"✅ Successfully created XLSX file: {output_file_path}")
        print(f"📊 Exported {len(data)} overrides with fields: {', '.join(headers)}")

        return True

    except Exception as e:
        print(f"❌ Failed to create XLSX file: {str(e)}")
        return False

def _auto_adjust_column_widths(sheet, headers, data):
    """
    Automatically adjust column widths based on content length
    """
    try:
        max_lengths = []

        # Initialize with header lengths
        for header in headers:
            max_lengths.append(len(str(header)))

        # Update with data cell lengths
        for item in data:
            # Extract values in the same order as headers
            row_values = [
                item.get('TagNumber', ''),
                item.get('Description', ''),
                item.get('OverrideType', {}).get('ShortForm') or item.get('OverrideType', {}).get('Title', ''),
                item.get('OverrideMethod', {}).get('ShortForm') or item.get('OverrideMethod', {}).get('Title', ''),
                item.get('OverrideAppliedState', {}).get('Title', ''),
                item.get('OverrideRemovedState', {}).get('Title', ''),
                item.get('Comment', ''),
                item.get('AdditionalValueAppliedState', ''),
                item.get('AdditionalValueRemovedState', ''),
                item.get('CurrentState', {}).get('Title', '')
            ]

            for col_index, cell_value in enumerate(row_values):
                if col_index < len(max_lengths):
                    cell_length = len(str(cell_value))
                    if cell_length > max_lengths[col_index]:
                        max_lengths[col_index] = cell_length

        # Apply calculated widths (starting from column A for data headers)
        for col_index, max_len in enumerate(max_lengths, 1):
            adjusted_width = min(max_len + 2, 50)
            column_letter = get_column_letter(col_index)
            sheet.column_dimensions[column_letter].width = adjusted_width

        print(f"📐 Adjusted {len(max_lengths)} column widths")

    except Exception as e:
        print(f"⚠️  Could not auto-adjust column widths: {str(e)}")

def main():
    """
    Main function to convert JSON to XLSX
    """
    # Input JSON file path
    json_file_path = "soc_resources\overrides_SOC_1797350.json"  # Change this to your JSON file path

    # Optional: specify output file path, or leave as None for auto-generation
    output_file_path = None  # This will auto-generate: soc_resources/SOC_FromJSON_overrides_TIMESTAMP.xlsx

    success = create_xlsx_from_json(json_file_path, output_file_path)

    if success:
        print("🎉 JSON to XLSX conversion completed successfully!")
    else:
        print("💥 JSON to XLSX conversion failed!")

if __name__ == "__main__":
    main()