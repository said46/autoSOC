# soc_exporter.py
from selenium.webdriver.common.by import By
from selenium.common.exceptions import (NoSuchElementException, TimeoutException)
from selenium.webdriver.support.wait import WebDriverWait
import logging
import configparser
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

from base_web_bot import BaseWebBot
from soc_base_mixin import SOC_BaseMixin


class SOC_Exporter(BaseWebBot, SOC_BaseMixin):
    """
    Specialized bot for exporting SOC overrides table to Excel.
    Exports data in the same format expected by the SOC_Importer.
    """

    def __init__(self, soc_id=None):
        BaseWebBot.__init__(self)
        SOC_BaseMixin.__init__(self)
        self.load_configuration()
        self.SOC_ID_PATTERN = r"^\d{7,8}$"
        self.SOC_id = soc_id if soc_id else None

    def set_soc_id(self, soc_id: str) -> None:
        self.SOC_id = soc_id

    @property
    def base_link(self) -> str:
        return self._base_link

    def load_configuration(self) -> None:
        config = configparser.ConfigParser(interpolation=None)
        config.read(self.config_file, encoding="utf8")

        self.user_name = config.get('Settings', 'user_name', fallback='xxxxxx')
        raw_password = config.get('Settings', 'password', fallback='******')
        self.password = self.process_password(raw_password)

        if '\n' in self.password:
            self.password = 'INCORRECT PASSWORD'

        self._base_link = config.get('Settings', 'base_link', fallback='http://eptw.sakhalinenergy.ru/')
        self.MAX_WAIT_PAGE_LOAD_DELAY_SECONDS = config.getint('Settings', 'MAX_WAIT_PAGE_LOAD_DELAY_SECONDS', fallback=20)

        logging.info(f"‚úÖ Configuration loaded from {self.config_file}")

    def check_if_overrides_exist(self) -> bool:
        try:
            overrides_section = self.driver.find_element(By.XPATH, "//label[contains(text(), '–û—Ç—á–µ—Ç –ø–æ —Ñ–æ—Ä—Å–∏—Ä–æ–≤–∫–∞–º')]")
            parent_panel = overrides_section.find_element(By.XPATH, "./ancestor::div[contains(@class, 'issow-panel')]")
            
            no_data_elements = parent_panel.find_elements(By.XPATH, ".//*[contains(text(), '–ù–µ—Ç –¥–∞–Ω–Ω—ã—Ö') or contains(text(), 'No data')]")
            if no_data_elements:
                logging.info("‚ÑπÔ∏è No overrides data found")
                return False
            return True

        except NoSuchElementException:
            logging.warning("‚ö†Ô∏è Overrides section not found")
            return False
        except Exception as e:
            logging.error(f"‚ùå Error checking overrides: {e}")
            return False

    def extract_overrides_table_data(self):
        """
        Extract SOC overrides data matching the importer's expected format.
        
        Expected column order for importer:
        TagNumber, Description, OverrideType, OverrideMethod, Comment, 
        AppliedState, AdditionalValueAppliedState, RemovedState, AdditionalValueRemovedState
        """
        try:
            logging.info("üîç Extracting SOC overrides data...")

            if not self.check_if_overrides_exist():
                return [], []

            # Wait for Kendo framework and grid data
            WebDriverWait(self.driver, self.MAX_WAIT_PAGE_LOAD_DELAY_SECONDS).until(
                lambda _: self.driver.execute_script("return typeof kendo !== 'undefined'")
            )

            WebDriverWait(self.driver, self.MAX_WAIT_PAGE_LOAD_DELAY_SECONDS).until(
                lambda _: self.driver.execute_script("""
                    var grid = $('#Overrides').data('kendoGrid');
                    return grid && grid.dataSource && grid.dataSource.data().length > 0;
                """)
            )

            script = """
            var grid = $('#Overrides').data('kendoGrid');
            if (!grid || !grid.dataSource) return null;

            var data = grid.dataSource.data();
            if (data.length === 0) return null;

            // Match the importer's expected column order (9 columns)
            var headers = [
                'TagNumber', 'Description', 'OverrideType', 'OverrideMethod', 'Comment',
                'AppliedState', 'AdditionalValueAppliedState', 'RemovedState', 'AdditionalValueRemovedState'
            ];

            var rows = data.map(item => [
                item.TagNumber || '',
                item.Description || '',
                item.OverrideType ? (item.OverrideType.Title || '') : '',
                item.OverrideMethod ? (item.OverrideMethod.Title || '') : '',
                item.Comment || '',
                item.OverrideAppliedState ? (item.OverrideAppliedState.Title || '') : '',
                item.AdditionalValueAppliedState || '',
                item.OverrideRemovedState ? (item.OverrideRemovedState.Title || '') : '',
                item.AdditionalValueRemovedState || ''
            ]);

            return {headers: headers, rows: rows};
            """

            result = self.driver.execute_script(script)
            if not result:
                return [], []

            headers = result.get('headers', [])
            rows = result.get('rows', [])

            if headers and rows:
                logging.info(f"‚úÖ Extracted {len(rows)} overrides with {len(headers)} columns")
                logging.info(f"üìä Columns: {', '.join(headers)}")
                return headers, rows

            return [], []

        except TimeoutException:
            logging.warning("‚ö†Ô∏è Timeout waiting for data")
            return [], []
        except Exception as e:
            logging.error(f"‚ùå Failed to extract data: {e}")
            return [], []

    def navigate_to_soc_details(self) -> None:
        try:
            soc_details_url = self.base_link + f"Soc/Details/{self.SOC_id}"
            logging.info(f"üåê Navigating to: {soc_details_url}")

            self.driver.get(soc_details_url)

            WebDriverWait(self.driver, self.MAX_WAIT_PAGE_LOAD_DELAY_SECONDS).until(
                lambda driver: driver.execute_script("return document.readyState") == "complete"
            )

            self.error_404_not_present_check()
            self.url_contains_SOC_Details_check()

            logging.info("‚úÖ Successfully navigated to SOC Details")
        except Exception as e:
            self.process_exception("‚ùå Navigation failed", e)

    def create_excel_file(self, headers: list, rows: list, filename: str = None) -> bool:
        """
        Create Excel file with fixed column order matching the importer's expectations.
        
        File structure:
        - Rows 1-4: Metadata
        - Row 6: Headers (fixed order for importer)
        - Rows 7+: Data rows
        """
        if not headers or not rows:
            logging.error("‚ùå No data to export")
            return False

        try:
            # Use fixed filename for importer compatibility
            if filename is None:
                #filename = f"soc_import_export/overrides.xlsx"
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"soc_import_export/SOC_{self.SOC_id}_overrides_{timestamp}.xlsx"                

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = f'SOC_{self.SOC_id}'

            # Add metadata
            metadata = [
                "SOC Overrides Export",
                f"SOC ID: {self.SOC_id}",
                f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                f"Total Overrides: {len(rows)}",
                "",  # Empty row for spacing
                # Row 6 will contain headers
            ]
            for i, value in enumerate(metadata, 1):
                sheet.cell(row=i, column=1, value=value)

            # Write headers at row 6 (matching importer expectation)
            header_row = 6
            for col_index, header in enumerate(headers, 1):
                cell = sheet.cell(row=header_row, column=col_index, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
                
                # Add light background color to headers
                cell.fill = openpyxl.styles.PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            # Write data starting from row 7 (matching importer expectation)
            for row_index, row_data in enumerate(rows, header_row + 1):
                for col_index, cell_value in enumerate(row_data, 1):
                    sheet.cell(row=row_index, column=col_index, value=cell_value)

            # Adjust column widths
            self._auto_adjust_column_widths(sheet, headers, rows)

            workbook.save(filename)
            logging.info(f"üíæ Saved to: {filename}")
            logging.info(f"üìÅ File ready for import with {len(rows)} overrides")
            return True

        except Exception as e:
            logging.error(f"‚ùå Failed to create Excel: {e}")
            return False

    def _auto_adjust_column_widths(self, sheet, headers: list, rows: list) -> None:
        """
        Auto-adjust column widths for better readability.
        """
        try:
            max_lengths = [len(str(header)) for header in headers]

            for row in rows:
                for col_index, cell_value in enumerate(row):
                    if col_index < len(max_lengths):
                        cell_length = len(str(cell_value))
                        if cell_length > max_lengths[col_index]:
                            max_lengths[col_index] = cell_length

            for col_index, max_len in enumerate(max_lengths, 1):
                adjusted_width = min(max_len + 2, 50)
                column_letter = get_column_letter(col_index)
                sheet.column_dimensions[column_letter].width = adjusted_width

        except Exception as e:
            logging.warning(f"‚ö†Ô∏è Could not adjust column widths: {e}")

    def extract_and_export_overrides(self) -> None:
        """
        Extract overrides and export to Excel in importer-compatible format.
        """
        try:
            headers, rows = self.extract_overrides_table_data()

            if not headers or not rows:
                msg = f"‚ÑπÔ∏è No overrides found for SOC {self.SOC_id}"
                logging.info(msg)
                self._inject_message_with_wait(msg, style_addons={'color': 'orange'})
            else:
                if self.create_excel_file(headers, rows):
                    msg = f"‚úÖ SOC {self.SOC_id} overrides exported successfully ({len(rows)} records)"
                    logging.info(msg)
                    self._inject_message_with_wait(msg, style_addons={'color': 'darkorange'})
                else:
                    msg = f"‚ùå Failed to save Excel for SOC {self.SOC_id}"
                    self._inject_message_with_wait(msg, style_addons={'color': 'red'})

        except Exception as e:
            logging.error(f"‚ùå Export error: {e}")
            self._inject_message_with_wait(f"‚ùå Export error: {e}", style_addons={'color': 'red'})

    def run(self, standalone=False):
        """
        Main execution workflow for SOC export.
        """
        try:
            logging.info("üöÄ Starting SOC export")

            if standalone:
                self.navigate_to_base()
                self.enter_credentials_and_prepare_soc_input()
                self.wait_for_soc_input_and_submit()

            self.navigate_to_soc_details()
            self.extract_and_export_overrides()

            logging.info("üèÅ SOC export completed")
            self._inject_message_with_wait("üèÅ SOC export completed", style_addons={'color': 'blue'})

        except Exception as e:
            logging.error(f"‚ùå SOC automation failed: {e}")
            try:
                self._inject_message_with_wait(f"‚ùå SOC automation failed: {e}", style_addons={'color': 'red'})
            except:
                self.safe_exit()


if __name__ == "__main__":
    bot = SOC_Exporter()
    bot.run(standalone=True)